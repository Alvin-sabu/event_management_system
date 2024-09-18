from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.contrib.auth.models import User
from django.contrib.auth.views import (
    PasswordResetView, PasswordResetDoneView, 
    PasswordResetConfirmView, PasswordResetCompleteView
)
from django.core.mail import EmailMessage, send_mail
from django.db import IntegrityError
from django.db.models import Count, F, Q, Value, FloatField
from django.db.models.functions import Coalesce, ExtractMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.safestring import mark_safe
from django.views.generic import DetailView

from calendar import HTMLCalendar
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from io import BytesIO
import csv
import json
import openpyxl
import random
import string
import zipfile
from openpyxl.utils import get_column_letter
from textblob import TextBlob

from .forms import (
    CustomUserCreationForm, DepartmentForm, EventForm, EventPhotoForm,
    EventRegistrationForm, FeedbackForm, FestForm, FrontPageVideoForm,
    NewsItemForm, ProfileForm, RegistrationForm, ReplyForm, SelectEventForm,
    SelectWinnersForm, SignUpForm, TeamMemberForm, UserUpdateForm,
    ProfileUpdateForm, CollegePosterForm, ContactMessageForm
)
from .models import (
    CollegePoster, ContactMessage, Department, Event, EventPhoto, Feedback,
    Fest, FrontPageVideo, NewsItem, Prize, Profile, Registration, TeamMember
)
from .notifications import notify_results_published
from .utils import admin_required, send_prize_notification_emails
from django.contrib.admin.views.decorators import staff_member_required


def index(request):
    login_form = AuthenticationForm()
    signup_form = SignUpForm()
    videos = FrontPageVideo.objects.all()  # Assuming FrontPageVideo is your model
    
    # Fetch the latest news items using created_at field
    news_items = NewsItem.objects.all().order_by('-created_at')[:3]  # Get the 3 most recent news items

    # Fetch all event photos
    event_photos = EventPhoto.objects.all()  # Assuming EventPhoto is your model

    return render(request, 'index.html', {
        'login_form': login_form,
        'signup_form': signup_form,
        'videos': videos,
        'news_items': news_items,
        'event_photos': event_photos,  # Pass event photos to the template
    })


def signup(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            try:
                user = form.save()
                login(request, user)  # Log the user in after successful signup
                return redirect('login')  # Redirect to login page or any other page
            except IntegrityError:  # Catch the unique constraint error
                form.add_error('email', 'This email is already in use.')
    else:
        form = CustomUserCreationForm()

    return render(request, 'signup.html', {'form': form})



def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                if user.is_staff:
                    messages.success(request, 'Admin login successful!')
                    return redirect('admin_dashboard')  # Redirect to admin dashboard
                else:
                    messages.success(request, 'You are now logged in!')
                    return redirect('event_list')  # Redirect to event list for regular users
            else:
                messages.error(request, 'Invalid username or password. Please try again.')
        else:
            messages.error(request, 'Form validation failed. Please check your input.')
    else:
        form = AuthenticationForm()

    return render(request, 'events/login.html', {'form': form})



class CustomPasswordResetView(PasswordResetView):
    template_name = 'custom_password_reset.html'
    success_url = reverse_lazy('password_reset_done')

class CustomPasswordResetDoneView(PasswordResetDoneView):
    template_name = 'custom_password_reset_done.html'

class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = 'custom_password_reset_confirm.html'
    success_url = reverse_lazy('password_reset_complete')

class CustomPasswordResetCompleteView(PasswordResetCompleteView):
    template_name = 'custom_password_reset_complete.html'

def custom_logout(request):
    logout(request)
    return redirect(reverse('index'))


def about(request):
    team_members = TeamMember.objects.all()
    for member in team_members:
        print(member.image.url)  # Use 'image' if the field is named 'image'
    return render(request, 'events/about.html', {'team_members': team_members})

def contact(request):
    return render(request, 'contact.html')


@login_required
def event_list(request):
    current_time = timezone.now()
    
    # Fetch all events
    all_events = Event.objects.all()
    
    upcoming_events = []
    current_events = []
    ended_events = []
    
    for event in all_events:
        # Convert event date and time to timezone-aware datetime
        event_start = timezone.make_aware(datetime.combine(event.date, event.time))
        event_end = event_start + timedelta(hours=1)  # Assuming events last 1 hour, adjust as needed
        
        if current_time < event.registration_start:
            upcoming_events.append(event)
        elif event_end <= current_time:
            ended_events.append(event)
        else:
            current_events.append(event)
    
    # Get posters
    posters = CollegePoster.objects.all()
    
    return render(request, 'events/event_list.html', {
        'upcoming_events': upcoming_events,
        'current_events': current_events,
        'ended_events': ended_events,
        'posters': posters,
        'events': all_events,
    })

# @login_required
# def event_list(request):
#     current_time = timezone.now()
    
#     # Filter events based on current time
#     upcoming_events = Event.objects.filter(registration_start__gt=current_time)
#     current_events = Event.objects.filter(registration_start__lte=current_time)

#     # Fetch posters for the template
#     posters = CollegePoster.objects.all()
    
#     # Render the template with the events and posters
#     return render(request, 'events/event_list.html', {
#         'upcoming_events': upcoming_events,
#         'current_events': current_events,
#         'posters': posters,
#         'events': Event.objects.all(),
#     })

# @login_required
# def event_list(request):
#     current_time = timezone.now()
    # 'events': Event.objects.all(),
#     upcoming_events = Event.objects.filter(registration_start__gt=current_time)
#     current_events = Event.objects.filter(registration_start__lte=current_time)
    
#     posters = CollegePoster.objects.all()  # Fetch posters for the template

#     return render(request, 'events/event_list.html', {
#         'upcoming_events': upcoming_events,
#         'current_events': current_events,
#         'posters': posters  # Add posters to the context
#     })

@login_required
def event_detail(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    registration_form = EventRegistrationForm(initial={'event_id': event_id})
    
    context = {
        'event': event,
        'registration_form': registration_form,
        'events': Event.objects.all(),
    }
    
    return render(request, 'events/event_detail.html', context)


@login_required
def registration_closed(request):
    return render(request, 'events/registration_closed.html')


@login_required
def already_registered(request):
    return render(request, 'events/already_registered.html')


@login_required
def register_for_event(request, event_id):
    event = get_object_or_404(Event, id=event_id)

    # Check if registration period has started
    if timezone.now() < event.registration_start:
        return redirect('registration_not_started')

    # Check if the user has already registered for this event
    if Registration.objects.filter(event=event, user=request.user).exists():
        return redirect('already_registered')

    # Check if event is full
    if event.is_full():
        return redirect('registration_closed')

    # Check if the current time is after the event start time
    current_time = timezone.now()
    event_datetime = datetime.combine(event.date, event.time)
    event_datetime = timezone.make_aware(event_datetime)
    if current_time > event_datetime:
        return redirect('registration_closed')

    if request.method == 'POST':
        form = EventRegistrationForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            token = generate_token()

            registration = Registration.objects.create(
                event=event,
                user=request.user,
                name=name,
                email=email,
                token=token
            )

            # Decrease the event capacity
            event.capacity -= 1
            event.save()

            # Prepare the email content
            email_subject = 'Event Registration Confirmation'
            email_body = render_to_string('emails/registration_confirmation.html', {'registration': registration})

            email_message = EmailMessage(
                subject=email_subject,
                body=email_body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email],
            )
            email_message.content_subtype = 'html'  # Set the content type to HTML
            email_message.send(fail_silently=False)

            return redirect('registration_detail', registration_id=registration.id)
    else:
        form = EventRegistrationForm()

    return render(request, 'events/register_for_event.html', {'event': event, 'form': form})



@login_required
def registration_detail(request, registration_id):
    registration = get_object_or_404(Registration, id=registration_id)
    return render(request, 'events/registration_detail.html', {'registration': registration})

def generate_token():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))


@login_required
def download_token(request, registration_id):
    registration = get_object_or_404(Registration, pk=registration_id, user=request.user)
    registration.generate_token_pdf()

    # Serve the PDF for download
    response = HttpResponse(registration.token_pdf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{registration.token}.pdf"'
    return response


@login_required
def user_registration_history(request):
    registrations = Registration.objects.filter(user=request.user)
    return render(request, 'events/user_registration_history.html', {'registrations': registrations})



class EventCalendar(HTMLCalendar):
    def __init__(self, events):
        super().__init__()
        self.events = self.group_by_day(events)

    def group_by_day(self, events):
        grouped_events = {}
        for event in events:
            day = event.date.day
            if day not in grouped_events:
                grouped_events[day] = []
            grouped_events[day].append(event)
        return grouped_events

    def formatday(self, day, weekday):
        events_from_day = self.events.get(day, [])
        events_html = '<ul>'
        for event in events_from_day:
            events_html += f'<li><a href="{reverse("event_detail", args=[event.id])}">{event.title}</a></li>'
        events_html += '</ul>'
        if day != 0:
            cssclass = self.cssclasses[weekday]
            if events_from_day:
                cssclass += ' event-day'
            return f"<td class='{cssclass}'><span class='day'>{day}</span> {events_html}</td>"
        return '<td></td>'

    def formatmonth(self, year, month, withyear=True):
        events = Event.objects.filter(date__year=year, date__month=month)
        self.events = self.group_by_day(events)
        return super().formatmonth(year, month, withyear)

    def get_previous_month(self, date):
        first_day = date.replace(day=1)
        prev_month = first_day - timedelta(days=1)
        return prev_month.year, prev_month.month

    def get_next_month(self, date):
        last_day = date.replace(day=28) + timedelta(days=4)  # This will always get the next month
        next_month = last_day.replace(day=1)
        return next_month.year, next_month.month


def calendar_view(request, year=datetime.now().year, month=datetime.now().strftime('%m')):
    year = int(year)
    month = int(month)
    date = datetime(year, month, 1)
    cal = EventCalendar(Event.objects.filter(date__year=year, date__month=month))
    html_cal = cal.formatmonth(year, month)
    prev_year, prev_month = cal.get_previous_month(date)
    next_year, next_month = cal.get_next_month(date)
    prev_url = reverse('calendar_view', kwargs={'year': prev_year, 'month': prev_month})
    next_url = reverse('calendar_view', kwargs={'year': next_year, 'month': next_month})
    context = {
        'calendar': mark_safe(html_cal),
        'prev_url': prev_url,
        'next_url': next_url,
        'year': year,
        'month': month,
    }
    return render(request, 'events/calendar.html', context)


def events_on_day(request, year, month, day):
    events = Event.objects.filter(date__year=year, date__month=month, date__day=day)
    events_list = [{'title': event.title, 'description': event.description, 'time': event.time.strftime('%H:%M')} for event in events]
    return JsonResponse(events_list, safe=False)

@login_required
def feedback(request, event_id):
    event = get_object_or_404(Event, id=event_id)

    # Check if the event has ended
    if not event.has_ended():
        return redirect('event_list')  # Redirect to event list if feedback is not available

    # Check if the user has registered for this event
    if not Registration.objects.filter(event=event, user=request.user).exists():
        return redirect('event_list')  # Redirect to event list if the user has not registered

    # Check if the user has already provided feedback for this event
    if Feedback.objects.filter(event=event, user=request.user).exists():
        return redirect('feedback_thanks')  # Redirect to thank you page if feedback is already provided

    if request.method == 'POST':
        form = FeedbackForm(request.POST)
        if form.is_valid():
            # Save the feedback
            feedback = Feedback(
                event=event,
                user=request.user,
                overall_rating=form.cleaned_data['overall_rating'],
                content_relevance=form.cleaned_data['content_relevance'],
                speaker_evaluation=form.cleaned_data['speaker_evaluation'],
                organization_rating=form.cleaned_data['organization_rating'],
                suggestions=form.cleaned_data['suggestions']
            )
            feedback.save()
            return redirect('feedback_thanks')  # Redirect after successful feedback submission
    else:
        form = FeedbackForm()

    return render(request, 'events/feedback.html', {'form': form, 'event': event})



def feedback_thanks(request):
    return render(request, 'events/feedback_thanks.html')


@login_required
def event_feedback_report(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    feedback_list = Feedback.objects.filter(event=event)

    # Calculate statistics (e.g., average ratings)
    ratings = {
        'overall_rating': {'Excellent': 0, 'Good': 0, 'Average': 0, 'Poor': 0},
        'content_relevance': {'Very relevant': 0, 'Somewhat relevant': 0, 'Neutral': 0, 'Not very relevant': 0, 'Not relevant at all': 0},
        'speaker_evaluation': {'Excellent': 0, 'Good': 0, 'Average': 0, 'Poor': 0},
        'organization_rating': {'Very well organized': 0, 'Well organized': 0, 'Neutral': 0, 'Poorly organized': 0, 'Very poorly organized': 0},
    }

    for feedback in feedback_list:
        ratings['overall_rating'][feedback.overall_rating] += 1
        ratings['content_relevance'][feedback.content_relevance] += 1
        ratings['speaker_evaluation'][feedback.speaker_evaluation] += 1
        ratings['organization_rating'][feedback.organization_rating] += 1

    context = {
        'event': event,
        'feedback_list': feedback_list,
        'ratings': ratings
    }
    return render(request, 'events/feedback_report.html', context)


def results_page(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    if not event.has_ended():
        return render(request, 'error.html', {'message': 'Results are not available for ongoing events.'})
    
    if not event.results_published:
        return render(request, 'error.html', {'message': 'Results have not been published yet.'})
    
    # Prefetching the related Prize objects for the registrations
    registrations = Registration.objects.filter(event=event).prefetch_related(
        Prefetch('prizes', queryset=Prize.objects.all())
    )
    
    # Sorting the registrations based on the prize type
    sorted_registrations = sorted(
        registrations, key=lambda r: r.prizes.first().prize_type if r.prizes.exists() else ''
    )
    
    return render(request, 'results_page.html', {
        'event': event,
        'registrations': sorted_registrations,
        'no_results': not sorted_registrations,
        'current_year': timezone.now().year
    })

def contact(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        message = request.POST.get('message')

        if name and email and message:
            ContactMessage.objects.create(name=name, email=email, message=message)
            send_mail(
                'Contact Form Submission',
                f'Name: {name}\nEmail: {email}\nMessage: {message}',
                email,
                ['alvinksabu200115@gmail.com']  # Replace with your admin email
            )
            return redirect('contact_success')
        else:
            # Handle the case where one or more fields are missing
            return render(request, 'contact.html', {'error': 'Please fill in all fields'})
    return render(request, 'contact.html')

def contact_success(request):
    return render(request, 'contact_success.html')


@staff_member_required
def contact_message_list(request):
    messages = ContactMessage.objects.all()
    return render(request, 'admin/contactmessage/contact_message_list.html', {'messages': messages})

@staff_member_required
def ContactMessageAdminView(request):
    contact_messages = ContactMessage.objects.all()
    return render(request, 'admin/contactmessage/changelist.html', {'contact_messages': contact_messages})


def upload_team_member(request):
    if request.method == 'POST':
        form = TeamMemberForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('about')
    else:
        form = TeamMemberForm()
    return render(request, 'upload_team_member.html', {'form': form})

@login_required
def profile_view(request):
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        profile = Profile(user=request.user)
        profile.save()
        
    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=request.user.profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            return redirect('profile')
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = ProfileUpdateForm(instance=request.user.profile)

    context = {
        'user_form': user_form,
        'profile_form': profile_form
    }
    return render(request, 'events/profile.html', context)

def create_event(request):
    # Your view logic here
    return render(request, 'events/create_event.html')

@login_required
def profile_edit(request):
    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=request.user.profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, 'Profile updated successfully')
            return redirect('profile')
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = ProfileUpdateForm(instance=request.user.profile)

    context = {
        'user_form': user_form,
        'profile_form': profile_form
    }
    return render(request, 'events/profile_edit.html', context)

from .models import CollegePoster

def home(request):
    posters = CollegePoster.objects.all()
    return render(request, 'event.html', {'posters': posters})


# Helper function to restrict access to admin users
def admin_required(user):
    return user.is_staff

@login_required
@user_passes_test(admin_required)
def custom_admin_logout(request):
    logout(request)
    return redirect('index')




def admin_required(user):
    return user.is_staff  # Modify this check based on your specific admin requirements



@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_dashboard(request):
    # Fetch necessary data for the dashboard
    events = Event.objects.all()
    departments = Department.objects.all()
    users = User.objects.all()
    fests = Fest.objects.all()
    news_items = NewsItem.objects.all()
    photos = EventPhoto.objects.all() 
    departments = Department.objects.prefetch_related('event_set__registrations').all()
    
    
     # Fetch all photos

    # Calculate totals
    total_events = events.count()
    total_departments = departments.count()
    total_users = users.count()
    total_prizes = events.aggregate(total_prizes=Count('prizes'))['total_prizes'] or 0

    # Event Performance Analysis
    events = events.annotate(total_registrations=Count('registrations'))
    for event in events:
        event.capacity_utilization = (event.total_registrations / event.capacity) * 100 if event.capacity else 0
    top_events = events.order_by('-total_registrations')[:5]

    # Department Contributions Analysis
    department_contributions = Department.objects.annotate(total_registrations=Count('event__registrations'))
    department_data = []
    for department in departments:
        events = department.event_set.all()
        event_data = []
        for event in events:
            registrations = event.registrations.all()
            event_data.append({
                'event': event,
                'registrations': registrations,
            })
        department_data.append({
            'department': department,
            'events': event_data,
        })

    # User Engagement Analysis
    active_users = Registration.objects.values('user__username').annotate(total=Count('id')).order_by('-total')[:10]
    repeat_participants = Registration.objects.values('user__username').annotate(events_participated=Count('event', distinct=True)).filter(events_participated__gt=1).order_by('-events_participated')

    # Feedback Analysis
    feedbacks = Feedback.objects.all()
    positive_feedback = feedbacks.filter(overall_rating__gte=4).count()
    neutral_feedback = feedbacks.filter(overall_rating__gte=3, overall_rating__lt=4).count()
    negative_feedback = feedbacks.filter(overall_rating__lt=3).count()

    # Data for charts
    current_year = datetime.now().year
    months = [f'{month:02d}' for month in range(1, 13)]
    month_labels = [datetime(2000, i, 1).strftime('%B') for i in range(1, 13)]
    registration_counts = [Registration.objects.filter(event__date__year=current_year, event__date__month=month).count() for month in range(1, 13)]

    # Fest Date Filtered Events
    fest_date_filtered_events = []
    if fests.exists():
        for fest in fests:
            events_in_fest_date_range = events.filter(date__range=(fest.start_date, fest.end_date))
            fest_date_filtered_events.append({
                'fest': fest.name,
                'events': events_in_fest_date_range
            })

    # Ended Events
    ended_events = events.filter(date__lt=timezone.now().date())

    # Recent Activities Helper Function
    def get_recent_activities():
        activities = []

        # Recent events created
        recent_events = events.order_by('-created_at')[:5]
        for event in recent_events:
            activities.append(f"New event created: {event.title} on {event.created_at.strftime('%Y-%m-%d') if event.created_at else 'unknown date'}")

        # Recent departments added
        recent_departments = departments.order_by('-created_at')[:5]
        for department in recent_departments:
            activities.append(f"Department added: {department.name} on {department.created_at.strftime('%Y-%m-%d') if department.created_at else 'unknown date'}")

        # Recent posters uploaded
        recent_posters = CollegePoster.objects.order_by('-created_at')[:5]
        for poster in recent_posters:
            activities.append(f"New college poster uploaded on {poster.created_at.strftime('%Y-%m-%d') if poster.created_at else 'unknown date'}")

        # Recent feedbacks exported
        recent_feedbacks = Feedback.objects.order_by('-exported_at')[:5]
        for feedback in recent_feedbacks:
            activities.append(f"Feedback exported for {feedback.event.title} on {feedback.exported_at.strftime('%Y-%m-%d') if feedback.exported_at else ''}")

        # Recent photos added
        recent_photos = photos.order_by('-created_at')[:5]
        for photo in recent_photos:
            activities.append(f"New photo added: {photo.description} on {photo.created_at.strftime('%Y-%m-%d') if photo.created_at else 'unknown date'}")

        return activities

    recent_activities = get_recent_activities()

    # Fetch all contact messages and count unread ones
    contact_messages = ContactMessage.objects.all().order_by('-created_at')
    unread_messages_count = contact_messages.filter(replied=False).count()

    # Pass all data to the template
    context = {
        'total_events': total_events,
        'total_departments': total_departments,
        'total_users': total_users,
        'total_prizes': total_prizes,
        'recent_activities': recent_activities,
        'top_events': top_events,
        'department_contributions': department_contributions,
        'active_users': active_users,
        'repeat_participants': repeat_participants,
        'positive_feedback': positive_feedback,
        'neutral_feedback': neutral_feedback,
        'negative_feedback': negative_feedback,
        'month_labels': json.dumps(month_labels),
        'registration_counts': json.dumps(registration_counts),
        'fest_date_filtered_events': fest_date_filtered_events,
        'ended_events': ended_events,
        'events': events,  # Include events for prize management links
        'contact_messages': contact_messages,  # Include all contact messages
        'unread_messages_count': unread_messages_count,  # Include unread messages count
        'news_items': news_items,
        'photos': photos, 
        'department_data': department_data, # Include photos for management links
    }

    return render(request, 'custom_admin/dashboard.html', context)




@login_required
@user_passes_test(lambda u: u.is_superuser)
def admin_messages(request):
    contact_messages = ContactMessage.objects.all().order_by('-created_at')
    unread_messages_count = contact_messages.filter(replied=False).count()

    # Mark all messages as replied (instead of read)
    ContactMessage.objects.filter(replied=False).update(replied=True)

    context = {
        'contact_messages': contact_messages,
        'unread_messages_count': unread_messages_count,
    }

    return render(request, 'custom_admin/messages.html', context)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def reply_to_message(request, message_id):
    message = get_object_or_404(ContactMessage, id=message_id)
    if request.method == 'POST':
        form = ReplyForm(request.POST)
        if form.is_valid():
            reply_message = form.cleaned_data['reply_message']
            subject = f'Re: {message.message[:30]}...'
            email_message = (
                f'Dear {message.name},\n\n'
                f'Thank you for reaching out to us. Here is our reply to your query:\n\n'
                f'Original Message: {message.message}\n\n'
                f'Reply: {reply_message}\n\n'
                'Best regards,\nEvent Hub,\nMarian College Kuttikkanam'
            )
            send_mail(
                subject,
                email_message,
                settings.DEFAULT_FROM_EMAIL,
                [message.email]
            )
            message.replied = True
            message.save()
            messages.success(request, 'Reply sent successfully!')
            return redirect('admin_messages')
    else:
        form = ReplyForm()
    
    return render(request, 'custom_admin/reply_to_message.html', {'form': form, 'message': message})


def get_fest_dates(request, fest_id):
    try:
        fest = Fest.objects.get(id=fest_id)
        return JsonResponse({
            'start_date': fest.start_date.strftime('%Y-%m-%d'),
            'end_date': fest.end_date.strftime('%Y-%m-%d')
        })
    except Fest.DoesNotExist:
        return JsonResponse({'error': 'Fest not found.'}, status=404)
    
@login_required
@user_passes_test(admin_required)
def create_event(request):
    if request.method == 'POST':
        form = EventForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('admin_dashboard')
    else:
        form = EventForm()
    return render(request, 'custom_admin/create_event.html', {'form': form})

@login_required
@user_passes_test(admin_required)
def update_event(request, pk):
    event = get_object_or_404(Event, pk=pk)
    if request.method == 'POST':
        form = EventForm(request.POST, instance=event)
        if form.is_valid():
            form.save()
            return redirect('list_events')  # Redirect to event list or another relevant page
    else:
        form = EventForm(instance=event)
    return render(request, 'custom_admin/update_event.html', {'form': form})

@login_required
@user_passes_test(admin_required)
def delete_event(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    if request.method == 'POST':
        event.delete()
        return redirect('admin_dashboard')
    return render(request, 'custom_admin/delete_event.html', {'object': event})

@login_required
@user_passes_test(admin_required)
def approve_event(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    if request.method == 'POST':
        if event.date <= timezone.now().date():
            event.status = 'completed'
            event.results_published = True
            event.save()
            # Notify results published (pseudo-function)
            send_mail(
                'Results Published',
                'The results for the event "{}" have been published.'.format(event.title),
                'alvinksabu200115@gmail.com',
                ['alvinksabu200115@gmail.com'],
                fail_silently=False,
            )
        else:
            event.status = 'approved'
            event.save()
        return redirect('admin_dashboard')
    return render(request, 'custom_admin/approve_event.html', {'event': event})

@login_required
@user_passes_test(admin_required)
def create_department(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('list_departments')
    else:
        form = DepartmentForm()
    return render(request, 'custom_admin/create_department.html', {'form': form})

@login_required
@user_passes_test(admin_required)
def update_department_view(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            return redirect('list_departments')  # Redirect to the list of departments or another relevant page
    else:
        form = DepartmentForm(instance=department)
    return render(request, 'custom_admin/update_department.html', {'form': form})

@login_required
@user_passes_test(admin_required)
def delete_department_view(request, department_id):
    department = get_object_or_404(Department, id=department_id)
    if request.method == 'POST':
        department.delete()
        return redirect('admin_dashboard')
    return render(request, 'custom_admin/delete_department.html', {'object': department})

@login_required
def manage_fests(request):
    # Add logic to retrieve necessary data for managing fests
    fests = Fest.objects.all()
    context = {
        'fests': fests,
    }
    return render(request, 'custom_admin/manage_fests.html', context)

@login_required
@user_passes_test(admin_required)
def create_fest(request):
    if request.method == 'POST':
        form = FestForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('admin_dashboard')
    else:
        form = FestForm()
    return render(request, 'custom_admin/create_fest.html', {'form': form})

@login_required
@user_passes_test(admin_required)
def update_fest_view(request, fest_id):
    fest = get_object_or_404(Fest, id=fest_id)
    if request.method == 'POST':
        form = FestForm(request.POST, instance=fest)
        if form.is_valid():
            form.save()
            return redirect('manage_fests')  # Redirect to the fest management page or another relevant page
    else:
        form = FestForm(instance=fest)
    return render(request, 'custom_admin/update_fest.html', {'form': form})

@login_required
@user_passes_test(admin_required)
def delete_fest_view(request, fest_id):
    fest = get_object_or_404(Fest, id=fest_id)
    if request.method == 'POST':
        fest.delete()
        return redirect('admin_dashboard')
    return render(request, 'custom_admin/delete_fest.html', {'object': fest})



@login_required
@user_passes_test(admin_required)
def view_feedbacks(request):
    feedbacks = Feedback.objects.all()
    return render(request, 'custom_admin/view_feedbacks.html', {'feedbacks': feedbacks})




@login_required
@user_passes_test(admin_required)
def view_feedbacks(request):
    feedbacks = Feedback.objects.all().select_related('event', 'user')

    # Handle search
    search_query = request.GET.get('search', '')
    if search_query:
        feedbacks = feedbacks.filter(
            Q(event__title__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )

    # Handle filters
    event_filter = request.GET.get('event', '')
    if event_filter:
        feedbacks = feedbacks.filter(event__title__icontains=event_filter)
    
    user_filter = request.GET.get('user', '')
    if user_filter:
        feedbacks = feedbacks.filter(user__username__icontains=user_filter)

    overall_rating_filter = request.GET.get('overall_rating', '')
    if overall_rating_filter:
        feedbacks = feedbacks.filter(overall_rating=overall_rating_filter)
    
    content_relevance_filter = request.GET.get('content_relevance', '')
    if content_relevance_filter:
        feedbacks = feedbacks.filter(content_relevance=content_relevance_filter)
    
    speaker_evaluation_filter = request.GET.get('speaker_evaluation', '')
    if speaker_evaluation_filter:
        feedbacks = feedbacks.filter(speaker_evaluation=speaker_evaluation_filter)
    
    organization_rating_filter = request.GET.get('organization_rating', '')
    if organization_rating_filter:
        feedbacks = feedbacks.filter(organization_rating=organization_rating_filter)

    # Aggregated data for chart
    overall_ratings_counts = feedbacks.values('overall_rating').annotate(count=Count('id')).order_by('overall_rating')
    content_relevance_counts = feedbacks.values('content_relevance').annotate(count=Count('id')).order_by('content_relevance')

    context = {
        'feedbacks': feedbacks,
        'events': Feedback.objects.values_list('event__title', flat=True).distinct(),
        'users': Feedback.objects.values_list('user__username', flat=True).distinct(),
        'overall_ratings': Feedback.objects.values_list('overall_rating', flat=True).distinct(),
        'content_relevance': Feedback.objects.values_list('content_relevance', flat=True).distinct(),
        'speaker_evaluation': Feedback.objects.values_list('speaker_evaluation', flat=True).distinct(),
        'organization_rating': Feedback.objects.values_list('organization_rating', flat=True).distinct(),
        'overall_ratings_counts': overall_ratings_counts,
        'content_relevance_counts': content_relevance_counts,
    }
    
    return render(request, 'custom_admin/view_feedbacks.html', context)

@login_required
@user_passes_test(admin_required)
def export_feedback(request):
    feedbacks = Feedback.objects.all()
    
    # Generate CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="feedbacks.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Feedback ID', 'Event', 'User', 'Overall Rating', 'Content Relevance', 'Speaker Evaluation', 'Organization Rating', 'Suggestions'])  # Header based on model fields
    
    for feedback in feedbacks:
        writer.writerow([
            feedback.id,
            feedback.event.title,
            feedback.user.username,
            feedback.overall_rating,
            feedback.content_relevance,
            feedback.speaker_evaluation,
            feedback.organization_rating,
            feedback.suggestions or ''  # Ensure suggestions are handled as optional
        ])  # Export fields based on model definition
    
    return response

@login_required
@user_passes_test(admin_required)
def manage_college_posters(request):
    posters = CollegePoster.objects.all()
    return render(request, 'custom_admin/manage_college_posters.html', {'posters': posters})

@login_required
@user_passes_test(admin_required)
def upload_college_poster(request):
    if request.method == 'POST':
        form = CollegePosterForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('manage_college_posters')
    else:
        form = CollegePosterForm()
    return render(request, 'custom_admin/upload_college_poster.html', {'form': form})

@login_required
@user_passes_test(admin_required)
def update_college_poster(request, poster_id):
    poster = get_object_or_404(CollegePoster, id=poster_id)
    if request.method == 'POST':
        form = CollegePosterForm(request.POST, request.FILES, instance=poster)
        if form.is_valid():
            form.save()
            return redirect('manage_college_posters')
    else:
        form = CollegePosterForm(instance=poster)
    return render(request, 'custom_admin/update_college_poster.html', {'form': form, 'poster': poster})

@login_required
@user_passes_test(admin_required)
def delete_college_poster(request, poster_id):
    poster = get_object_or_404(CollegePoster, id=poster_id)
    if request.method == 'POST':
        poster.delete()
        return redirect('manage_college_posters')
    return render(request, 'custom_admin/delete_college_poster.html', {'object': poster})


def list_events(request):
    # Get filter and search query from request
    filter_value = request.GET.get('filter', '')
    search_query = request.GET.get('search', '')

    # Initial queryset for events
    events = Event.objects.all()

    # Filter events by department or day based on the combined filter value
    if filter_value:
        if filter_value.startswith('department-'):
            department_id = filter_value.split('-')[1]
            events = events.filter(department_id=department_id)
        elif filter_value.startswith('day-'):
            day = filter_value.split('-')[1]
            # Fetch the current fest to use its date range
            now = timezone.now()
            current_fest = Fest.objects.filter(start_date__lte=now, end_date__gte=now).first()
            if current_fest:
                fest_start_date = current_fest.start_date
                start_date = fest_start_date + timezone.timedelta(days=int(day) - 1)
                end_date = start_date + timezone.timedelta(days=1)
                events = events.filter(date__range=(start_date, end_date))
    
    # Further filter events by search query if provided
    if search_query:
        events = events.filter(title__icontains=search_query)

    # Define top events (e.g., based on capacity or any other criteria)
    top_events = events.filter(capacity__gte=100)  # Example criterion for top events

    # Fetch all departments for the filter dropdown
    departments = Department.objects.all()

    # Fetch days for filtering dropdown (assuming the fest is ongoing)
    days = []
    now = timezone.now()
    current_fest = Fest.objects.filter(start_date__lte=now, end_date__gte=now).first()
    if current_fest:
        fest_start_date = current_fest.start_date
        total_days = (current_fest.end_date - fest_start_date).days + 1
        days = list(range(1, total_days + 1))

    # Group events by the day of the fest
    events_by_day = defaultdict(list)
    if current_fest:
        fest_start_date = current_fest.start_date
        for event in events:
            event_day = (event.date - fest_start_date).days + 1
            if 1 <= event_day <= (current_fest.end_date - fest_start_date).days + 1:
                events_by_day[event_day].append(event)

    context = {
        'events_by_day': dict(events_by_day),  # Converting defaultdict to regular dict for the template
        'top_events': top_events,
        'departments': departments,
        'days': days,
    }

    return render(request, 'custom_admin/list_events.html', context)

def list_departments(request):
    departments = Department.objects.all()
    print(departments)  # Debugging line
    return render(request, 'custom_admin/list_departments.html', {'departments': departments})


def list_fests(request):
    fests = Fest.objects.all()
    return render(request, 'custom_admin/list_fests.html', {'fests': fests})

def list_college_posters(request):
    posters = CollegePoster.objects.all()
    return render(request, 'custom_admin/list_college_posters.html', {'posters': posters})


from django.shortcuts import render
from django.db.models import Count, F, FloatField, Value
from django.db.models.functions import Coalesce
from events.models import Event
from django.shortcuts import render
from django.db.models import Count, F, FloatField, Value
from django.db.models.functions import Coalesce
from events.models import Event  # Replace with your actual model import

def event_performance_analysis(request):
    # Total registrations per event
    events = Event.objects.annotate(
        total_registrations=Count('registrations')
    )

    # Capacity utilization (avoid division by zero by using Coalesce)
    events = events.annotate(
        capacity_utilization=Coalesce(
            (F('total_registrations') * 100.0 / F('capacity')),
            Value(0),
            output_field=FloatField()
        )
    )

    # Top events by registration
    top_events = events.order_by('-total_registrations')[:5]

    # Department contribution
    department_contributions = Event.objects.values('department__name').annotate(
        total=Count('registrations')
    ).order_by('-total')

    context = {
        'events': events,
        'top_events': top_events,
        'department_contributions': department_contributions,
    }
    return render(request, 'custom_admin/event_performance_analysis.html', context)



from django.shortcuts import render
from django.db.models import Count
from events.models import Registration, Feedback
from textblob import TextBlob

def analyze_sentiment(feedback_text):
    analysis = TextBlob(feedback_text)
    if analysis.sentiment.polarity > 0:
        return 'positive'
    elif analysis.sentiment.polarity == 0:
        return 'neutral'
    else:
        return 'negative'

def user_engagement_analysis(request):
    # Active users (users with the most registrations)
    active_users = Registration.objects.values('user__username').annotate(
        total=Count('id')
    ).order_by('-total')[:10]

    # Repeat participants (users who participated in more than one event)
    repeat_participants = Registration.objects.values('user__username').annotate(
        events_participated=Count('event', distinct=True)
    ).filter(events_participated__gt=1).order_by('-events_participated')[:10]

    # Feedback analysis (sentiment counts)
    feedbacks = Feedback.objects.all()
    positive_feedback = sum(1 for feedback in feedbacks if analyze_sentiment(feedback.suggestions) == 'positive')
    neutral_feedback = sum(1 for feedback in feedbacks if analyze_sentiment(feedback.suggestions) == 'neutral')
    negative_feedback = sum(1 for feedback in feedbacks if analyze_sentiment(feedback.suggestions) == 'negative')

    # Event-wise feedback analysis
    event_feedbacks = Feedback.objects.values('event__title').annotate(
        feedback_count=Count('id')
    ).order_by('-feedback_count')

    context = {
        'active_users': active_users,
        'repeat_participants': repeat_participants,
        'positive_feedback': positive_feedback,
        'neutral_feedback': neutral_feedback,
        'negative_feedback': negative_feedback,
        'event_feedbacks': event_feedbacks,
    }
    return render(request, 'custom_admin/user_engagement_analysis.html', context)


@login_required
@user_passes_test(admin_required)
def manage_event_and_winners(request, event_id=None):
    selected_event = None
    winners_selected = False
    event_form = SelectEventForm(request.POST or None)
    winners_form = None

    if event_id:
        selected_event = get_object_or_404(Event, id=event_id)
        winners_selected = Prize.objects.filter(event=selected_event).exists()
        winners_form = SelectWinnersForm(event=selected_event)

        # Check if the event has ended (after 11:59 PM of the event date)
        event_end_time = timezone.make_aware(
            datetime.combine(selected_event.date, time(23, 59, 59)),
            timezone.get_current_timezone()
        )
        if timezone.now() < event_end_time:
            messages.error(request, "Winners can only be selected after the event has ended.")
            return redirect('manage_event_and_winners')

    if request.method == 'POST':
        if 'select_event' in request.POST:
            if event_form.is_valid():
                event_id = event_form.cleaned_data['event'].id
                return redirect('manage_event_and_winners', event_id=event_id)
        elif 'manage_winners' in request.POST:
            winners_form = SelectWinnersForm(request.POST, event=selected_event)
            if winners_form.is_valid():
                # Ensure the event has ended before allowing winners to be selected
                if timezone.now() < event_end_time:
                    messages.error(request, "Winners can only be selected after the event has ended.")
                    return redirect('manage_event_and_winners', event_id=event_id)

                Prize.objects.filter(event=selected_event).delete()
                winners = {}
                for prize_type in ['first_prize', 'second_prize', 'third_prize']:
                    registration = winners_form.cleaned_data.get(prize_type)
                    if registration:
                        Prize.objects.create(
                            event=selected_event,
                            registration=registration,
                            prize_type=prize_type.split('_')[0]
                        )
                        winners[prize_type] = registration

                selected_event.results_published = True
                selected_event.save()
                
                try:
                    send_prize_notification_emails(winners, selected_event)
                    messages.success(request, "Winners have been selected and notified successfully!")
                except Exception as e:
                    messages.error(request, f"Winners selected but email notification failed: {str(e)}")
                return redirect('manage_event_and_winners', event_id=event_id)
            else:
                messages.error(request, "Form validation failed. Please correct the errors.")

    context = {
        'event_form': event_form,
        'winners_form': winners_form,
        'selected_event': selected_event,
        'winners_selected': winners_selected
    }
    return render(request, 'custom_admin/manage_event_and_winners.html', context)

def send_prize_notification_emails(winners, event):
    for prize_type, registration in winners.items():
        if registration:
            user = registration.user
            subject = f'Congratulations! You have won the {prize_type.replace("_", " ").title()} Prize'
            message = f'Hello {user.first_name},\n\nCongratulations! You have been awarded the {prize_type.replace("_", " ").title()} Prize for the event "{event.title}".'
            from_email = settings.DEFAULT_FROM_EMAIL
            recipient_list = [user.email]
            
            try:
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=from_email,
                    recipient_list=recipient_list,
                    fail_silently=False,
                )
                print(f"Email sent successfully to {user.email}")
            except Exception as e:
                print(f"Failed to send email to {user.email}: {str(e)}")
                raise  # Re-raise the exception to be caught in the calling function


# Team Member Views
def list_team_members(request):
    team_members = TeamMember.objects.all()
    return render(request, 'custom_admin/list_team_members.html', {'team_members': team_members})

def add_team_member(request):
    if request.method == 'POST':
        form = TeamMemberForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('list_team_members')
    else:
        form = TeamMemberForm()
    return render(request, 'custom_admin/add_team_member.html', {'form': form})

def edit_team_member(request, pk):
    team_member = get_object_or_404(TeamMember, pk=pk)
    if request.method == 'POST':
        form = TeamMemberForm(request.POST, request.FILES, instance=team_member)
        if form.is_valid():
            form.save()
            return redirect('list_team_members')
    else:
        form = TeamMemberForm(instance=team_member)
    return render(request, 'custom_admin/edit_team_member.html', {'form': form})

def delete_team_member(request, pk):
    team_member = get_object_or_404(TeamMember, pk=pk)
    team_member.delete()
    return redirect('list_team_members')

# Front Page Video Views
def list_front_page_videos(request):
    videos = FrontPageVideo.objects.all()
    return render(request, 'custom_admin/list_front_page_videos.html', {'videos': videos})

def add_front_page_video(request):
    if request.method == 'POST':
        form = FrontPageVideoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('list_front_page_videos')
    else:
        form = FrontPageVideoForm()
    return render(request, 'custom_admin/add_front_page_video.html', {'form': form})

def edit_front_page_video(request, pk):
    video = get_object_or_404(FrontPageVideo, pk=pk)
    if request.method == 'POST':
        form = FrontPageVideoForm(request.POST, request.FILES, instance=video)
        if form.is_valid():
            form.save()
            return redirect('list_front_page_videos')
    else:
        form = FrontPageVideoForm(instance=video)
    return render(request, 'custom_admin/edit_front_page_video.html', {'form': form})

def delete_front_page_video(request, pk):
    video = get_object_or_404(FrontPageVideo, pk=pk)
    video.delete()
    return redirect('list_front_page_videos')

def media_management(request):
    return render(request, 'custom_admin/media_management.html')



def event_registration_list(request, event_id):
    # Retrieve the event object or return a 404 error if not found
    event = get_object_or_404(Event, id=event_id)
    
    # Retrieve all registrations for the event using the 'registrations' related name
    registrations = event.registrations.all()
    
    # Context to pass to the template
    context = {
        'event': event,
        'registrations': registrations
    }
    
    # Render the template with the context data
    return render(request, 'custom_admin/event_registration_list.html', context)

@login_required
@user_passes_test(lambda u: u.is_superuser)
def add_news_item(request):
    if request.method == 'POST':
        form = NewsItemForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_dashboard')
    else:
        form = NewsItemForm()
    return render(request, 'custom_admin/add_news_item.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def update_news_item(request, id):
    news_item = get_object_or_404(NewsItem, id=id)
    if request.method == 'POST':
        form = NewsItemForm(request.POST, request.FILES, instance=news_item)
        if form.is_valid():
            form.save()
            return redirect('admin_dashboard')
    else:
        form = NewsItemForm(instance=news_item)
    return render(request, 'custom_admin/update_news_item.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_news_item(request, id):
    news_item = get_object_or_404(NewsItem, id=id)
    if request.method == 'POST':
        news_item.delete()
        return redirect('admin_dashboard')
    return render(request, 'custom_admin/delete_news_item.html', {'news_item': news_item})


class NewsDetailView(DetailView):
    model = NewsItem
    template_name = 'events/news_detail.html'
    context_object_name = 'news_item'


class NewsDetailView(DetailView):
    model = NewsItem
    template_name = 'events/news_detail.html'
    context_object_name = 'news_item'

def event_statistics(request):
    # Query database for event registration counts by month
    registrations = Registration.objects.filter(
        registration_date__year=2024  # Adjust year or remove filter as needed
    ).annotate(
        month=ExtractMonth('registration_date')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')

    # Prepare data for Chart.js
    month_labels = [f"Month {i+1}" for i in range(12)]
    registration_counts = [0] * 12

    for registration in registrations:
        month = registration['month'] - 1  # Adjust for zero-based index
        registration_counts[month] = registration['count']

    context = {
        'month_labels': month_labels,
        'registration_counts': registration_counts,
    }

    return render(request, 'custom/admin_dashboard.html', context)


@login_required
def add_photo(request):
    if request.method == 'POST':
        form = EventPhotoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('index')  # Redirect to a relevant page after adding
    else:
        form = EventPhotoForm()
    return render(request, 'custom_admin/add_photo.html', {'form': form})

@login_required
def edit_photo(request, photo_id):
    photo = get_object_or_404(EventPhoto, id=photo_id)
    if request.method == 'POST':
        form = EventPhotoForm(request.POST, request.FILES, instance=photo)
        if form.is_valid():
            form.save()
            return redirect('index')  # Redirect to a relevant page after editing
    else:
        form = EventPhotoForm(instance=photo)
    return render(request, 'custom_admin/edit_photo.html', {'form': form, 'photo': photo})

@login_required
def delete_photo(request, photo_id):
    photo = get_object_or_404(EventPhoto, id=photo_id)
    if request.method == 'POST':
        photo.delete()
        return redirect('index')  # Redirect to a relevant page after deleting
    return render(request, 'custom_admin/confirm_delete.html', {'photo': photo})

@login_required
def list_photos(request):
    photos = EventPhoto.objects.all()
    return render(request, 'custom_admin/list_photos.html', {'photos': photos})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def download_department_wise_registrations(request):
    # Fetch departments and related event registrations
    departments = Department.objects.prefetch_related('event_set__registrations').all()

    # Create an in-memory output file for the ZIP archive
    zip_output = BytesIO()

    # Create a ZIP file
    with zipfile.ZipFile(zip_output, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for department in departments:
            # Sanitize department name for folder name
            sanitized_department_name = department.name.replace('/', '_').replace('\\', '_')
            department_folder = f"{sanitized_department_name}/"

            # Ensure department folder is created in the ZIP file
            zip_file.writestr(department_folder, b'')  # Create an empty folder

            for event in department.event_set.all():
                # Create a workbook for the current event
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Registrations Report"

                # Write header row
                headers = ['Participant Name', 'Email', 'Registration Date', 'Token']
                for col_num, header in enumerate(headers, 1):
                    column_letter = get_column_letter(col_num)
                    sheet[f'{column_letter}1'] = header

                # Write event-wise registration data
                row_num = 2  # Start writing data from row 2 (after the header)
                for registration in event.registrations.all():
                    print(f"Writing to Excel: Name: {registration.name}, Email: {registration.email}, Token: {registration.token}")  # Debug line
                    sheet[f'A{row_num}'] = registration.name
                    sheet[f'B{row_num}'] = registration.email
                    sheet[f'C{row_num}'] = registration.registration_date.strftime('%Y-%m-%d %H:%M:%S')
                    sheet[f'D{row_num}'] = registration.token  # Add token to the Excel file
                    row_num += 1

                # Save the workbook to a BytesIO object
                event_output = BytesIO()
                workbook.save(event_output)
                event_output.seek(0)

                # Add the event workbook to the department folder in the ZIP file
                sanitized_event_title = event.title.replace('/', '_').replace('\\', '_')
                event_filename = f"{sanitized_event_title}_registrations.xlsx"
                zip_file.writestr(department_folder + event_filename, event_output.getvalue())

    # Set the cursor of the in-memory file to the beginning
    zip_output.seek(0)

    # Create a response with the correct content type for ZIP files
    response = HttpResponse(
        zip_output,
        content_type='application/zip'
    )
    response['Content-Disposition'] = 'attachment; filename="department_registrations.zip"'

    return response


@login_required
@user_passes_test(lambda u: u.is_superuser)
def view_registrations(request):
    # Fetch departments and related event registrations
    departments = Department.objects.prefetch_related('event_set__registrations').all()
    
    if request.GET.get('download') == 'true':
        # Create an in-memory output file for the ZIP archive
        zip_output = BytesIO()

        # Create a ZIP file
        with zipfile.ZipFile(zip_output, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for department in departments:
                # Create a folder for the department
                department_folder = f"{department.name}/"
                
                for event in department.event_set.all():
                    # Create a workbook for the current event
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    sheet.title = "Registrations Report"

                    # Write header row
                    headers = ['Participant Name', 'Email', 'Registration Date', 'Token']
                    for col_num, header in enumerate(headers, 1):
                        column_letter = get_column_letter(col_num)
                        sheet[f'{column_letter}1'] = header

                    # Write event-wise registration data
                    row_num = 2  # Start writing data from row 2 (after the header)
                    for registration in event.registrations.all():
                        sheet[f'A{row_num}'] = registration.name
                        sheet[f'B{row_num}'] = registration.email
                        sheet[f'C{row_num}'] = registration.registration_date.strftime('%Y-%m-%d %H:%M:%S')
                        sheet[f'D{row_num}'] = registration.token  # Add token to the Excel file
                        row_num += 1

                    # Save the workbook to a BytesIO object
                    event_output = BytesIO()
                    workbook.save(event_output)
                    event_output.seek(0)

                    # Add the event workbook to the department folder in the ZIP file
                    event_filename = f"{event.title}_registrations.xlsx"
                    zip_file.writestr(department_folder + event_filename, event_output.getvalue())

        # Set the cursor of the in-memory file to the beginning
        zip_output.seek(0)

        # Create a response with the correct content type for ZIP files
        response = HttpResponse(
            zip_output,
            content_type='application/zip'
        )
        response['Content-Disposition'] = 'attachment; filename="department_registrations.zip"'

        return response

    # Render the template with departments
    context = {
        'departments': departments
    }
    return render(request, 'custom_admin/view_registrations.html', context)

